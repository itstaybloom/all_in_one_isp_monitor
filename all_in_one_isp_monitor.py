import subprocess
import platform
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re
import os
import pandas as pd
import streamlit as st
import plotly.express as px

# =========================
# CONFIG
# =========================
LATENCY_WARNING_THRESHOLD = 100
REPORT_FOLDER = "ping_reports"
IP_LIST_PATH = "ip_list.txt"

if not os.path.exists(REPORT_FOLDER):
    os.makedirs(REPORT_FOLDER)

param = "-n" if platform.system().lower() == "windows" else "-c"

# =========================
# PING FUNCTIONS
# =========================
def get_isp(ip):
    try:
        url = f"http://ip-api.com/json/{ip}"
        response = requests.get(url, timeout=5).json()
        return response.get("isp", "Unknown")
    except:
        return "Lookup Failed"

def ping_ip(ip):
    try:
        command = ["ping", param, "2", ip]
        output = subprocess.check_output(command, universal_newlines=True)
        latency = "N/A"
        for line in output.splitlines():
            match = re.search(r"time[=<]([\d\.]+)", line)
            if match:
                latency = match.group(1)
                break
        return "PASS", latency
    except:
        return "FAIL", "N/A"

def generate_report():
    with open(IP_LIST_PATH, "r") as file:
        ip_list = [ip.strip() for ip in file if ip.strip()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ping Report"
    ws.append(["IP Address", "ISP Name", "Status", "Latency (ms)", "Timestamp"])

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    pass_count = 0
    fail_count = 0
    high_latency_count = 0
    any_failure = False

    for ip in ip_list:
        status, latency = ping_ip(ip)
        isp = get_isp(ip)
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([ip, isp, status, latency, timestamp])
        row = ws.max_row

        if status == "PASS":
            pass_count += 1
        else:
            fail_count += 1
            any_failure = True
            for cell in ws[row]:
                cell.fill = red_fill

        if latency != "N/A":
            try:
                if float(latency) > LATENCY_WARNING_THRESHOLD:
                    high_latency_count += 1
                    ws.cell(row=row, column=4).fill = yellow_fill
            except:
                pass

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 3

    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Total IPs Checked", len(ip_list)])
    summary_ws.append(["Total PASS", pass_count])
    summary_ws.append(["Total FAIL", fail_count])
    summary_ws.append(["High Latency (>100ms)", high_latency_count])
    summary_ws.append(["Report Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    filename = os.path.join(REPORT_FOLDER, f"ping_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filename)

    return filename, any_failure

# =========================
# STREAMLIT DASHBOARD
# =========================
st.set_page_config(page_title="ISP Monitoring Dashboard", layout="wide")
st.markdown("<meta http-equiv='refresh' content='60'>", unsafe_allow_html=True)
st.title("🌐 Multi-ISP Monitoring Dashboard")

# Run Ping Test
if st.button("▶ Run Ping Test & Generate New Report"):
    file_created, any_failure = generate_report()
    st.success(f"New Report Created: {file_created}")
    if any_failure:
        st.warning("⚠️ Some IPs have FAILED the ping test! Check the report for details.")

# Load latest report safely
files = sorted([os.path.join(REPORT_FOLDER, f) for f in os.listdir(REPORT_FOLDER) if f.endswith(".xlsx")], reverse=True)
if not files:
    st.warning("No reports found yet. Click the button above to generate one.")
    st.stop()

latest_file = files[0]
try:
    latest_df = pd.read_excel(latest_file, sheet_name="Ping Report")
    latest_summary = pd.read_excel(latest_file, sheet_name="Summary")
except Exception as e:
    st.error(f"Failed to read Excel file: {e}")
    st.stop()

def get_summary_value(df, metric_name):
    val = df.loc[df["Metric"] == metric_name, "Value"]
    return int(val.values[0]) if not val.empty else 0

# Summary Metrics
col1, col2, col3, col4 = st.columns(4)
col1.metric("Total IPs", get_summary_value(latest_summary, "Total IPs Checked"))
col2.metric("PASS", get_summary_value(latest_summary, "Total PASS"))
col3.metric("FAIL", get_summary_value(latest_summary, "Total FAIL"))
col4.metric("High Latency", get_summary_value(latest_summary, "High Latency (>100ms)"))

# Filter by ISP
st.subheader("Filter by ISP")
isp_list = ["All"] + sorted(latest_df["ISP Name"].astype(str).unique().tolist())
selected_isp = st.selectbox("Select ISP", isp_list)

df_live = latest_df.copy()
if selected_isp != "All":
    df_live = df_live[df_live["ISP Name"] == selected_isp]

st.subheader("Live Status Table")
st.dataframe(df_live, use_container_width=True)

st.subheader("Failed Links (Live)")
st.dataframe(df_live[df_live["Status"]=="FAIL"], use_container_width=True)

# =========================
# Live ISP Health Chart
# =========================
st.subheader("ISP-wise Live Health (Stacked)")

df_live["High Latency"] = df_live["Latency (ms)"].apply(lambda x: 1 if x != "N/A" and float(x) > LATENCY_WARNING_THRESHOLD else 0)
isp_summary = df_live.groupby("ISP Name").agg({
    "PASS": lambda x: sum(x=="PASS"),
    "FAIL": lambda x: sum(x=="FAIL"),
    "High Latency": "sum"
}).reset_index()

isp_summary_melt = isp_summary.melt(id_vars="ISP Name", value_vars=["PASS","FAIL","High Latency"],
                                    var_name="Status", value_name="Count")
color_map = {"PASS":"green", "FAIL":"red", "High Latency":"yellow"}

fig = px.bar(isp_summary_melt, x="ISP Name", y="Count", color="Status",
             color_discrete_map=color_map, barmode="stack", text="Count")
fig.update_layout(yaxis_title="Number of IPs", xaxis_title="ISP Name")
st.plotly_chart(fig, use_container_width=True)

# =========================
# Historical ISP Stacked Chart
# =========================
st.subheader("📊 Historical ISP Trend (Stacked)")

history_frames = []
for file in files[::-1]:
    try:
        df_ping = pd.read_excel(file, sheet_name="Ping Report")
        timestamp = pd.to_datetime(pd.read_excel(file, sheet_name="Summary").loc[
            lambda x: x["Metric"]=="Report Generated At","Value"].values[0]
        )
        df_group = df_ping.groupby(["ISP Name","Status"]).size().unstack(fill_value=0)
        df_group["High Latency"] = df_ping["Latency (ms)"].apply(
            lambda x: 1 if x != "N/A" and float(x) > LATENCY_WARNING_THRESHOLD else 0
        )
        df_group["Timestamp"] = timestamp
        history_frames.append(df_group.reset_index())
    except:
        pass

if history_frames:
    history_df = pd.concat(history_frames, ignore_index=True)
    history_melt = history_df.melt(id_vars=["ISP Name","Timestamp"], value_vars=["PASS","FAIL","High Latency"],
                                   var_name="Status", value_name="Count")
    fig_hist = px.bar(history_melt, x="Timestamp", y="Count", color="Status",
                      color_discrete_map=color_map, barmode="stack", facet_col="ISP Name",
                      facet_col_wrap=2, text="Count")
    fig_hist.update_layout(yaxis_title="Number of IPs", xaxis_title="Timestamp")
    st.plotly_chart(fig_hist, use_container_width=True)

st.caption(f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
